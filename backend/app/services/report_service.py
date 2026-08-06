import csv
import io

from sqlalchemy.orm import Session

from app.models.dns import DNSLookup
from app.models.hash import FileHash
from app.models.jwt import JwtInspection
from app.models.password import PasswordCheck
from app.models.ssl import SSLCheck
from app.models.user import User
from app.models.whois import WhoisCheck


def _all_data(db: Session, user: User) -> dict:
    """Reúne todos los análisis del usuario en una estructura de datos."""
    return {
        "dns": (
            db.query(DNSLookup)
            .filter(DNSLookup.user_id == user.id)
            .order_by(DNSLookup.created_at.desc())
            .all()
        ),
        "ssl": (
            db.query(SSLCheck)
            .filter(SSLCheck.user_id == user.id)
            .order_by(SSLCheck.created_at.desc())
            .all()
        ),
        "whois": (
            db.query(WhoisCheck)
            .filter(WhoisCheck.user_id == user.id)
            .order_by(WhoisCheck.created_at.desc())
            .all()
        ),
        "passwords": (
            db.query(PasswordCheck)
            .filter(PasswordCheck.user_id == user.id)
            .order_by(PasswordCheck.created_at.desc())
            .all()
        ),
        "hashes": (
            db.query(FileHash)
            .filter(FileHash.user_id == user.id)
            .order_by(FileHash.created_at.desc())
            .all()
        ),
        "jwt": (
            db.query(JwtInspection)
            .filter(JwtInspection.user_id == user.id)
            .order_by(JwtInspection.created_at.desc())
            .all()
        ),
    }


def _format_dt(value) -> str:
    if not value:
        return "—"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value).replace("T", " ")[:16]


def build_csv(db: Session, user: User) -> bytes:
    data = _all_data(db, user)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Seccion", "Clave", "Valor"])

    for item in data["ssl"]:
        writer.writerow(["SSL", "dominio", item.domain])
        writer.writerow(["SSL", "valido", item.is_valid])
        writer.writerow(["SSL", "expira_en_dias", item.days_left])
        writer.writerow(["SSL", "expira", _format_dt(item.expires_at)])

    for item in data["whois"]:
        writer.writerow(["WHOIS", "dominio", item.domain])
        writer.writerow(["WHOIS", "registrador", item.registrar])
        writer.writerow(["WHOIS", "creado", _format_dt(item.creation_date)])
        writer.writerow(["WHOIS", "expira", _format_dt(item.expiration_date)])

    for item in data["dns"]:
        writer.writerow(["DNS", "dominio", item.domain])
        for record in item.records:
            writer.writerow(["DNS", f"registro_{record.type}", record.value])

    for item in data["passwords"]:
        writer.writerow(["PASSWORD", "longitud", item.password_length])
        writer.writerow(["PASSWORD", "entropia_bits", f"{item.entropy_bits:.1f}"])
        writer.writerow(["PASSWORD", "score", f"{item.score}/4"])
        writer.writerow(["PASSWORD", "tiempo_fuerza_bruta_seg", item.crack_time_seconds])

    for item in data["hashes"]:
        writer.writerow(["HASH", "archivo", item.filename])
        writer.writerow(["HASH", "sha256", item.sha256])
        writer.writerow(["HASH", "md5", item.md5])

    for item in data["jwt"]:
        writer.writerow(["JWT", "algoritmo", item.algorithm])
        writer.writerow(["JWT", "sujeto", item.subject])
        writer.writerow(["JWT", "advertencias", item.warnings])

    return buf.getvalue().encode("utf-8-sig")


def build_excel(db: Session, user: User) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = _all_data(db, user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    header_font = Font(bold=True)

    def write_table(sheet, title, headers, rows):
        sheet.append([])
        sheet.append([title])
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True, size=12)
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = header_font
        for row in rows:
            sheet.append(row)

    rows = []
    for item in data["ssl"]:
        rows.append(["SSL", item.domain, item.is_valid, item.days_left, _format_dt(item.expires_at)])
    for item in data["whois"]:
        rows.append(["WHOIS", item.domain, item.registrar, _format_dt(item.creation_date), _format_dt(item.expiration_date)])
    for item in data["dns"]:
        records = "; ".join(f"{r.type} {r.value}" for r in item.records)
        rows.append(["DNS", item.domain, records, "", ""])
    for item in data["passwords"]:
        rows.append(["PASSWORD", f"{item.password_length} chars", f"{item.entropy_bits:.1f} bits", f"{item.score}/4", f"{item.crack_time_seconds:.0f}s"])
    for item in data["hashes"]:
        rows.append(["HASH", item.filename, item.sha256, item.md5, ""])
    for item in data["jwt"]:
        rows.append(["JWT", item.algorithm, item.subject, "", ""])

    write_table(ws, "Todos los analisis de SentinelX", ["Modulo", "Campo 1", "Campo 2", "Campo 3", "Campo 4"], rows)

    for col in "ABCDE":
        ws.column_dimensions[col].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(db: Session, user: User) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    data = _all_data(db, user)
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 15 * mm
    y = height - margin

    def draw_title(text):
        nonlocal y
        c.setFont("Helvetica-Bold", 14)
        c.setFillColorRGB(0.1, 0.25, 0.55)
        c.drawString(margin, y, text)
        y -= 7 * mm

    def draw_line(label, value):
        nonlocal y
        if y < margin + 15 * mm:
            c.showPage()
            y = height - margin
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(margin, y, f"{label}:")
        c.drawRightString(width - margin, y, str(value)[:80])
        y -= 5 * mm

    draw_title(f"Informe de seguridad SentinelX - {user.email}")
    c.setFont("Helvetica", 9)
    c.drawString(margin, y, "Generado con datos reales de la cuenta.")
    y -= 5 * mm

    draw_title("Certificados SSL")
    if not data["ssl"]:
        draw_line("", "Sin datos")
    for item in data["ssl"]:
        draw_line("Dominio", item.domain)
        draw_line("Valido", "Si" if item.is_valid else "No")
        draw_line("Expira en dias", item.days_left)

    draw_title("WHOIS")
    if not data["whois"]:
        draw_line("", "Sin datos")
    for item in data["whois"]:
        draw_line("Dominio", item.domain)
        draw_line("Registrador", item.registrar)

    draw_title("DNS")
    if not data["dns"]:
        draw_line("", "Sin datos")
    for item in data["dns"]:
        draw_line("Dominio", item.domain)
        draw_line("Registros", "; ".join(f"{r.type} {r.value}" for r in item.records)[:80])

    draw_title("Contrasenas")
    if not data["passwords"]:
        draw_line("", "Sin datos")
    for item in data["passwords"]:
        draw_line("Longitud", item.password_length)
        draw_line("Entropia (bits)", f"{item.entropy_bits:.1f}")
        draw_line("Score", f"{item.score}/4")

    draw_title("Hashes de archivos")
    if not data["hashes"]:
        draw_line("", "Sin datos")
    for item in data["hashes"]:
        draw_line("Archivo", item.filename)
        draw_line("SHA-256", item.sha256)

    draw_title("JWT inspeccionados")
    if not data["jwt"]:
        draw_line("", "Sin datos")
    for item in data["jwt"]:
        draw_line("Algoritmo", item.algorithm)

    c.save()
    return buf.getvalue()
